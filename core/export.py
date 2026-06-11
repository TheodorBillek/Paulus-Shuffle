from __future__ import annotations
import csv
import io
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


def _build_rows(
    sessions: List[dict],
    students_by_id: Dict[str, dict],
    seats_by_id: Dict[str, dict],
) -> List[list]:
    rows = []
    for session in sessions:
        date = session.get("date", "")
        label = session.get("label", "") or date
        assignment: Dict[str, Optional[str]] = session.get("assignment", {})
        solo_set = set(session.get("solo_students", []))

        reverse = {v: k for k, v in assignment.items() if v is not None}

        tables_map: Dict[tuple, Dict[str, str]] = {}
        for seat in seats_by_id.values():
            key = (seat["row_idx"], seat["col_idx"])
            sid_at = reverse.get(seat["id"])
            if sid_at:
                tables_map.setdefault(key, {})[seat["side"]] = sid_at

        partner_map: Dict[str, Optional[str]] = {}
        for sides in tables_map.values():
            l, r = sides.get("L"), sides.get("R")
            if l and r:
                partner_map[l] = r
                partner_map[r] = l

        for student_id, seat_id in assignment.items():
            student = students_by_id.get(student_id)
            if not student:
                continue
            seat = seats_by_id.get(seat_id) if seat_id else None
            partner_id = partner_map.get(student_id)
            partner = students_by_id.get(partner_id) if partner_id else None

            rows.append([
                date,
                label,
                student["name"],
                student.get("gender", "X"),
                (seat["row_idx"] + 1) if seat else "",
                (seat["col_idx"] + 1) if seat else "",
                seat["side"] if seat else "",
                partner["name"] if partner else ("(solo)" if student_id in solo_set else ""),
                "Yes" if student_id in solo_set else "No",
            ])
    return rows


HEADERS = ["Date", "Session", "Student", "Gender", "Row", "Col", "Side", "Partner", "Solo"]


def sessions_to_xlsx(
    sessions: List[dict],
    students: List[dict],
    seats: List[dict],
) -> bytes:
    students_by_id = {s["id"]: s for s in students}
    seats_by_id = {s["id"]: s for s in seats}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Session History"

    ws.append(HEADERS)
    teal_fill = PatternFill("solid", fgColor="0F766E")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = teal_fill
        cell.alignment = Alignment(horizontal="center")

    for row in _build_rows(sessions, students_by_id, seats_by_id):
        ws.append(row)

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def sessions_to_csv(
    sessions: List[dict],
    students: List[dict],
    seats: List[dict],
) -> str:
    students_by_id = {s["id"]: s for s in students}
    seats_by_id = {s["id"]: s for s in seats}

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(HEADERS)
    for row in _build_rows(sessions, students_by_id, seats_by_id):
        writer.writerow(row)
    return output.getvalue()
